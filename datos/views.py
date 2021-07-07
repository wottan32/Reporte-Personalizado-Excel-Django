from django.shortcuts import render
from .models import Datos
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def home(request):
    queryset = Datos.objects.all()
    return render(request, 'index.html', {'queryset': queryset})


class ReportePersonalizadoExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        campo = int(request.GET.get(''))
        query = Datos.objects.filter(id=campo)
        wb = Workbook()
        bandera = True
        cont = 1
        controlador = 4
        for q in query:
            if bandera:
                ws = wb.active
                ws.title = 'Hoja' + str(cont)
                bandera = False
            else:
                ws = wb.create_sheet('Hoja' + str(cont))
            # Crear el título en la hoja
            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B1'].fill = PatternFill(start_color='66FFCC', end_color='66FFCC', fill_type="solid")
            ws['B1'].font = Font(name='Calibri', size=12, bold=True)
            ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'

            # Cambiar caracteristicas de las celdas
            ws.merge_cells('B1:N1')

            ws.row_dimensions[1].height = 25

            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 20
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 20
            ws.column_dimensions['K'].width = 20
            ws.column_dimensions['L'].width = 20
            ws.column_dimensions['M'].width = 20
            ws.column_dimensions['N'].width = 20
            # ws.column_dimensions['O'].width = 20
            # ws.column_dimensions['P'].width = 20
            # ws.column_dimensions['Q'].width = 20

            # Crear la cabecera
            ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['B3'].font = Font(name='Calibro', size=10, bold=True)
            ws['B3'] = 'empresa_id'

            ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['C3'].font = Font(name='Calibro', size=10, bold=True)
            ws['C3'] = 'nombre'

            ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['D3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['D3'].font = Font(name='Calibro', size=10, bold=True)
            ws['D3'] = 'razon_social'

            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['E3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['E3'].font = Font(name='Calibro', size=10, bold=True)
            ws['E3'] = 'rut'

            ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['F3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['F3'].font = Font(name='Calibro', size=10, bold=True)
            ws['F3'] = 'plazo_pago'

            ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['G3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['G3'].font = Font(name='Calibro', size=10, bold=True)
            ws['G3'] = 'oc'

            ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['H3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['H3'].font = Font(name='Calibro', size=10, bold=True)
            ws['H3'] = 'giro'

            ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['I3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['I3'].font = Font(name='Calibro', size=10, bold=True)
            ws['I3'] = 'contacto_factura'

            ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['J3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['J3'].font = Font(name='Calibro', size=10, bold=True)
            ws['J3'] = 'direccion_legal'

            ws['K3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['K3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['K3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['K3'].font = Font(name='Calibro', size=10, bold=True)
            ws['K3'] = 'comuna_legal'

            ws['L3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['L3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['L3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['L3'].font = Font(name='Calibro', size=10, bold=True)
            ws['L3'] = 'contactos'

            ws['M3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['M3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['M3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['M3'].font = Font(name='Calibro', size=10, bold=True)
            ws['M3'] = 'created'

            ws['N3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['N3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['N3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['N3'].font = Font(name='Calibro', size=10, bold=True)
            ws['N3'] = 'modified'

            # Pintamos los datos en el reporte
            ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=2).value = q.id

            ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=3).value = q.empresa_id

            ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=4).value = q.nombre

            ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=5).value = q.razon_social

            ws.cell(row=controlador, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=6).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=6).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=6).value = q.rut

            ws.cell(row=controlador, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=7).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=7).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=7).value = q.plazo_pago

            ws.cell(row=controlador, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=8).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=8).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=8).value = q.oc

            ws.cell(row=controlador, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=9).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=9).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=9).value = q.giro

            ws.cell(row=controlador, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=10).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=10).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=10).value = q.contacto_factura

            ws.cell(row=controlador, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=11).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=11).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=11).value = q.direccion_legal

            ws.cell(row=controlador, column=12).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=12).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=12).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=12).value = q.comuna_legal

            ws.cell(row=controlador, column=13).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=13).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=13).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=13).value = q.contactos

            ws.cell(row=controlador, column=14).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=14).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=14).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=14).value = q.created

            ws.cell(row=controlador, column=15).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=15).border = Border(left=Side(border_style="thin"),
                                                                right=Side(border_style="thin"),
                                                                top=Side(border_style="thin"),
                                                                bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=15).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=15).value = q.modified

            cont += 1

        # Establecer el nombre de mi archivo
        nombre_archivo = "ReportePersonalizadoExcel.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
